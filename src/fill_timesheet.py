from pathlib import Path as _Path2
import pandas as _pd
from pathlib import Path as _Path
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
import yaml

# PDF抽出～変形関数のインポート
from src.suica.Suica_pymupdf import extract_suica_history_pymupdf, add_year_to_dates
from src.suica.suica_transform import transform_commute
from src.suica.date_extractor import extract_history_date_pymupdf

# テンプレート・出力設定
TEMPLATE_PATH = Path('sample/d54ff476ff529c75ab262cbbed599019.xlsx')
OUTPUT_DIR = Path('output')
# 対象Suica PDF
PDF_PATH = Path('sample/JE80F121120754077_20231016_20240115160118.pdf')


def prepare_suica_df(pdf_path: Path) -> pd.DataFrame:
    report_date = extract_history_date_pymupdf(pdf_path)
    df_raw = extract_suica_history_pymupdf(pdf_path)
    df_with_year = add_year_to_dates(df_raw, report_date)
    df = transform_commute(df_with_year)
    df['日付'] = pd.to_datetime(df['日付'])
    df['year'] = df['日付'].dt.year
    df['month'] = df['日付'].dt.month
    return df


# 勤務表シートへのレポート日付書き込み

def write_report_date(ws, year: int, month: int):
    target = ws['D3']
    target.value = datetime(year, month, 1)
    target.number_format = "yyyy年m月"


# 通勤情報の書き込み（既存値がある場合は上書きしない）

def write_commute_entries(ws, group: pd.DataFrame):
    for d, route, fare in zip(group['日付'], group['通勤経路'], group['往復金額']):
        day = d.day
        row = day + 5
        cell_route = ws.cell(row, 22)
        if cell_route.value in (None, ''):
            cell_route.value = route
        cell_fare = ws.cell(row, 33)
        if cell_fare.value in (None, ''):
            cell_fare.value = fare


# デフォルト設定の読み込み

def load_defaults(config_path: Path) -> dict:
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


# 固定情報（所属支社、社員ID、氏名）の書き込み

def write_static_entries(ws, settings: dict):
    mapping = {
        'J3': 'branch',
        'Q3': 'employee_id',
        'Z3': 'name',
    }
    for cell_ref, key in mapping.items():
        cell = ws[cell_ref]
        if cell.value in (None, ''):
            cell.value = settings.get(key)


def make_timesheets(pdf_path: Path, template_path: Path, output_dir: Path):
    """
    Suica PDF から通勤情報を抽出し、Excelテンプレートに反映、月ごとに出力する。
    """
    output_dir.mkdir(exist_ok=True)
    df = prepare_suica_df(pdf_path)
    config_path = Path('setting/defaults.yaml')
    settings = load_defaults(config_path)
    for (year, month), group in df.groupby(['year', 'month']):
        wb = openpyxl.load_workbook(template_path)
        ws = wb['勤務表']
        write_report_date(ws, year, month)
        write_static_entries(ws, settings)
        write_commute_entries(ws, group)
        out_file = output_dir / f"勤務表_{year:04d}-{month:02d}.xlsx"
        wb.save(out_file)
        print(f"✓ 出力完了: {out_file}")


# GUI用: 抽出したSuica履歴をプレビューする


def preview_suica_records(pdf_path: _Path) -> _pd.DataFrame:
    """
    Suica PDFから全履歴を抽出し、日付をdatetime型に変換して返す（GUIプレビュー用）。
    """
    report_date = extract_history_date_pymupdf(pdf_path)
    df_raw = extract_suica_history_pymupdf(pdf_path)
    df_with_year = add_year_to_dates(df_raw, report_date)
    df_with_year['日付'] = _pd.to_datetime(df_with_year['日付'])
    return df_with_year


# GUI用: 選択されたSuica履歴から勤務表を生成する


def make_timesheets_from_records(
    df_records: _pd.DataFrame,
    template_path: _Path2 = TEMPLATE_PATH,
    output_dir: _Path2 = OUTPUT_DIR,
    config_path: _Path2 = Path('setting/defaults.yaml')
):
    """
    フィルタ済みのSuica履歴(df_records)から通勤情報を生成し、テンプレートに反映して勤務表を出力する。
    """
    output_dir.mkdir(exist_ok=True)
    # 通勤情報に変換
    df_comm = transform_commute(df_records)
    df_comm['日付'] = _pd.to_datetime(df_comm['日付'])
    df_comm['year'] = df_comm['日付'].dt.year
    df_comm['month'] = df_comm['日付'].dt.month
    # 設定読み込み
    settings = load_defaults(config_path)
    for (year, month), group in df_comm.groupby(['year', 'month']):
        wb = openpyxl.load_workbook(template_path)
        ws = wb['勤務表']
        write_report_date(ws, year, month)
        write_static_entries(ws, settings)
        write_commute_entries(ws, group)
        out_file = output_dir / f"勤務表_{year:04d}-{month:02d}.xlsx"
        wb.save(out_file)
        print(f"✓ 出力完了: {out_file}")
