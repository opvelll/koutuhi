import os
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import yaml

# PDF抽出～変形関数のインポート
from src.suica.Suica_pymupdf import extract_suica_history_pymupdf, add_year_to_dates
from src.suica.suica_transform import transform_commute
from src.suica.date_extractor import extract_history_date_pymupdf

# テンプレート・出力設定
TEMPLATE_PATH = Path('sample/d54ff476ff529c75ab262cbbed599019.xlsx')
OUTPUT_DIR = Path('output')
# 対象Suica PDF
PDF_PATH = Path('sample/JE80F121120754077_20231016_20240115160118.pdf')


# ヘルパ関数: Suicaデータの抽出と整形
def prepare_suica_df(pdf_path: Path) -> pd.DataFrame:
    report_date = extract_history_date_pymupdf(pdf_path)
    df_raw = extract_suica_history_pymupdf(pdf_path)
    df_with_year = add_year_to_dates(df_raw, report_date)
    df = transform_commute(df_with_year)
    df['日付'] = pd.to_datetime(df['日付'])
    df['year'] = df['日付'].dt.year
    df['month'] = df['日付'].dt.month
    return df


# ヘルパ関数: 勤務表シートへのレポート日付書き込み
def write_report_date(ws, year: int, month: int):
    target = ws['D3']
    target.value = datetime(year, month, 1)
    target.number_format = "yyyy年m月"


# ヘルパ関数: 通勤情報の書き込み（既存値がある場合は上書きしない）
def write_commute_entries(ws, group: pd.DataFrame):
    for d, route, fare in zip(group['日付'], group['通勤経路'], group['往復金額']):
        day = d.day
        row = day + 5
        cell_route = ws.cell(row, 22)
        if cell_route.value in (None, ''):
            cell_route.value = route
        cell_fare = ws.cell(row, 33)
        if cell_fare.value in (None, ''):
            cell_fare.value = fare


# ヘルパ関数: デフォルト設定の読み込み
def load_defaults(config_path: Path) -> dict:
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


# ヘルパ関数: 固定情報（所属支社、社員ID、氏名）の書き込み（既存値がある場合は上書きしない）
def write_static_entries(ws, settings: dict):
    # セル参照 -> (設定キー)
    mapping = {
        'J3': 'branch',      # 所属支社
        'Q3': 'employee_id',  # 社員ID
        'Z3': 'name',        # 氏名
    }
    for cell_ref, key in mapping.items():
        cell = ws[cell_ref]
        if cell.value in (None, ''):
            cell.value = settings.get(key)


def make_timesheets(pdf_path: Path, template_path: Path, output_dir: Path):
    """
    Suica PDF から通勤情報を抽出し、Excelテンプレートに反映、月ごとに出力する。
    """
    # 出力ディレクトリ作成
    output_dir.mkdir(exist_ok=True)

    # Suica履歴抽出
    df = prepare_suica_df(pdf_path)

    # 設定読み込み
    config_path = Path('setting/defaults.yaml')
    settings = load_defaults(config_path)
    for (year, month), group in df.groupby(['year', 'month']):
        # テンプレート読み込み
        wb = openpyxl.load_workbook(template_path)
        ws = wb['勤務表']
        # 年/月/1 形式の文字列を設定
        write_report_date(ws, year, month)
        # 固定情報の書き込み (J3, Q3, Z3)
        write_static_entries(ws, settings)

        # 日付ごとのデータを直接書き込み（日付セルを元に行を決定）
        write_commute_entries(ws, group)

        # ファイル出力
        out_file = output_dir / f"勤務表_{year:04d}-{month:02d}.xlsx"
        wb.save(out_file)
        print(f"✓ 出力完了: {out_file}")


# 確認関数: 出力ファイルの静的情報(J3, Q3, Z3)を検証
def verify_timesheets(output_dir: Path, config_path: Path):
    settings = load_defaults(config_path)
    for file in output_dir.glob('勤務表_*.xlsx'):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb['勤務表']
        assert ws['J3'].value == settings[
            'branch'], f"{file.name}: J3 expected {settings['branch']}, got {ws['J3'].value}"
        assert ws['Q3'].value == settings[
            'employee_id'], f"{file.name}: Q3 expected {settings['employee_id']}, got {ws['Q3'].value}"
        assert ws['Z3'].value == settings['name'], f"{file.name}: Z3 expected {settings['name']}, got {ws['Z3'].value}"
    print("✓ verification passed: all static entries are correct")


if __name__ == '__main__':
    make_timesheets(PDF_PATH, TEMPLATE_PATH, OUTPUT_DIR)
    # 出力結果の検証
    verify_timesheets(OUTPUT_DIR, Path('setting/defaults.yaml'))

    # 以下サンプル実行結果例
    # ✓ 出力完了: output\勤務表_2023-10.xlsx
    # ✓ 出力完了: output\勤務表_2023-11.xlsx
    # ✓ 出力完了: output\勤務表_2023-12.xlsx
    # ✓ 出力完了: output\勤務表_2024-01.xlsx
