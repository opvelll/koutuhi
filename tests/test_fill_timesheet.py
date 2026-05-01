from pathlib import Path

import openpyxl
import pandas as pd
import yaml

from src.fill_timesheet import make_timesheets_from_records


def _create_template(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '勤務表'
    wb.save(path)


def test_make_timesheets_from_records_writes_static_and_commute_entries(tmp_path):
    template_path = tmp_path / 'template.xlsx'
    config_path = tmp_path / 'defaults.yaml'
    output_dir = tmp_path / 'nested' / 'output'
    _create_template(template_path)
    config_path.write_text(
        yaml.safe_dump(
            {
                'branch': '東京',
                'employee_id': '12345',
                'name': '太郎 誠',
            },
            allow_unicode=True,
        ),
        encoding='utf-8',
    )

    records = pd.DataFrame(
        [
            {
                '日付': '2024/01/19',
                '種別1': '入',
                '利用駅1': '竹ノ塚',
                '種別2': '出',
                '利用駅2': '東武押上',
                '支払額': -261,
                '残額': 2233,
            },
            {
                '日付': '2024/01/19',
                '種別1': '入',
                '利用駅1': '東武押上',
                '種別2': '出',
                '利用駅2': '竹ノ塚',
                '支払額': -261,
                '残額': 1972,
            },
        ]
    )

    make_timesheets_from_records(records, template_path, output_dir, config_path)

    out_file = output_dir / '勤務表_2024-01.xlsx'
    wb = openpyxl.load_workbook(out_file, data_only=True)
    ws = wb['勤務表']
    assert ws['J3'].value == '東京'
    assert ws['Q3'].value == '12345'
    assert ws['Z3'].value == '太郎 誠'
    assert ws.cell(24, 22).value == '竹ノ塚～東武押上'
    assert ws.cell(24, 33).value == 522
